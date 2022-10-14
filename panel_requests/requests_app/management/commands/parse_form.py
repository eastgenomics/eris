#!usr/bin/env python

"""
This script (called in seed.py) will define functions to import panel
data from a request form and parse it for insertion into the database.
"""


import pandas as pd
from openpyxl import load_workbook


class FormParser:

    def __init__(self, filepath):
        self.filepath = filepath

    def get_form_data(self, fp):
        """ Pull in data from the request form file and parse into 4
        pandas DFs - general request info, panels info, genes, and
        regions.

        args:
            fp [str]: path to request form

        returns:
            general_info [dict]
            panel_df [pandas df]
            gene_df [pandas df]
            region_df [pandas df]
        """

        wb = load_workbook(filename=fp)
        ws = wb.active

        # identify row boundaries of each data table
        # this assumes there is -exactly- one blank row between each table

        row_count = 1

        for row in ws.iter_rows(max_col=1, values_only=True):
            if row[0] == 'Current panels':
                p_start = row_count + 1

            elif row[0] == 'Gene symbol':
                g_start = row_count + 1
                p_end = row_count - 2

            elif row[0] == 'Region name':
                r_start = row_count + 1
                g_end = row_count - 2

            elif row[0] == 'END OF FORM':
                r_end = row_count - 6
                break

            row_count += 1

        # create dict of general request info

        general_info = {
            'req_date': ws['B1'].value,
            'requester': ws['B2'].value,
            'ci': ws['B3'].value,
            'form_created_date': ws['B4'].value}

        # create df from panels info

        panel_df = pd.DataFrame({
            'names':
                [cell[0].value for cell in ws[f'A{p_start}': f'A{p_end}']],
            'sources':
                [cell[0].value for cell in ws[f'B{p_start}': f'B{p_end}']],
            'ext_ids':
                [cell[0].value for cell in ws[f'C{p_start}': f'C{p_end}']],
            'ext_versions':
                [cell[0].value for cell in ws[f'D{p_start}': f'D{p_end}']]}
                )

        # create df from genes info

        gene_df = pd.DataFrame({
            'hgncs':
                [cell[0].value for cell in ws[f'B{g_start}': f'B{g_end}']],
            'reasons':
                [cell[0].value for cell in ws[f'C{g_start}': f'C{g_end}']],
            'transcripts':
                [cell[0].value for cell in ws[f'D{g_start}': f'D{g_end}']],
            'trans_reasons':
                [cell[0].value for cell in ws[f'E{g_start}': f'E{g_end}']],
            'confs':
                [cell[0].value for cell in ws[f'F{g_start}': f'F{g_end}']],
            'pens':
                [cell[0].value for cell in ws[f'G{g_start}': f'G{g_end}']],
            'mops':
                [cell[0].value for cell in ws[f'H{g_start}': f'H{g_end}']],
            'mois':
                [cell[0].value for cell in ws[f'I{g_start}': f'I{g_end}']]})

        # create df from regions info

        region_df = pd.DataFrame({
            'names':
                [cell[0].value for cell in ws[f'A{r_start}': f'A{r_end}']],
            'chroms':
                [cell[0].value for cell in ws[f'B{r_start}': f'B{r_end}']],
            'starts_37':
                [cell[0].value for cell in ws[f'C{r_start}': f'C{r_end}']],
            'ends_37':
                [cell[0].value for cell in ws[f'D{r_start}': f'D{r_end}']],
            'starts_38':
                [cell[0].value for cell in ws[f'E{r_start}': f'E{r_end}']],
            'ends_38':
                [cell[0].value for cell in ws[f'F{r_start}': f'F{r_end}']],
            'reasons':
                [cell[0].value for cell in ws[f'G{r_start}': f'G{r_end}']],
            'confs':
                [cell[0].value for cell in ws[f'H{r_start}': f'H{r_end}']],
            'pens':
                [cell[0].value for cell in ws[f'I{r_start}': f'I{r_end}']],
            'mops':
                [cell[0].value for cell in ws[f'J{r_start}': f'J{r_end}']],
            'mois':
                [cell[0].value for cell in ws[f'K{r_start}': f'K{r_end}']],
            'types':
                [cell[0].value for cell in ws[f'L{r_start}': f'L{r_end}']],
            'var_types':
                [cell[0].value for cell in ws[f'M{r_start}': f'M{r_end}']],
            'haplos':
                [cell[0].value for cell in ws[f'N{r_start}': f'N{r_end}']],
            'triplos':
                [cell[0].value for cell in ws[f'O{r_start}': f'O{r_end}']],
            'overlaps':
                [cell[0].value for cell in ws[f'P{r_start}': f'P{r_end}']]})

        return general_info, panel_df, gene_df, region_df

    def setup_output_dict(self, ci, req_date, panel_df):
        """ Initialise a dict to hold relevant panel information.

        args:
            ci [str]: clinical indication to link the panel to
            req_date [str]: date of new panel request
            panel_df [pandas df]: PA panels new panel is based on

        returns:
            info_dict [dict]: initial dict of core panel info
        """

        # waiting for scientist input on what they want human-readable
        # panel names to look like

        # we also haven't decided what the id and version should look like
        # for non-panelapp panels

        info_dict = {
            'ci': ci,
            'req_date': req_date,
            'panel_source': 'Request',
            'panel_name': f'{ci}_request_{req_date}',
            'external_id': 'PLACEHOLDER',
            'panel_version': 'PLACEHOLDER',
            'genes': [],
            'regions': []}

        return info_dict

    def parse_genes(self, info_dict, gene_df):
        """ Iterate over every gene in the panel and retrieve the data
        needed to populate panel_gene and associated models. Only use
        genes with 'confidence_level' == '3'; i.e. 'green' genes.

        args:
            info_dict [dict]: holds data needed to populate db models
            gene_df [pandas df]: genes to include in new panel

        returns:
            info_dict [dict]: updated with gene data
        """

        for index, row in gene_df.iterrows():

            if row['hgncs']:

                gene_dict = {
                    'hgnc_id': row['hgncs'],
                    'gene_justification': row['reasons'],
                    'confidence_level': row['confs'],
                    'mode_of_inheritance': row['mois'],
                    'mode_of_pathogenicity': row['mops'],
                    'penetrance': row['pens'],
                    'transcript': row['transcripts'],
                    'transcript_justification': row['trans_reasons']}

                info_dict['genes'].append(gene_dict)

        return info_dict

    def parse_regions(self, info_dict, region_df):
        """ Iterate over every region in the panel and retrieve the data
        needed to populate panel_region and associated models. Only use
        regions with 'confidence_level' == '3'; i.e. 'green' regions.

        args:
            info_dict [dict]: holds data needed to populate db models
            region_df [pandas df]: regions to include in new panel

        returns:
            info_dict [dict]: updated with region data
        """

        for index, row in region_df.iterrows():

            if row['chroms']:

                region_dict = {
                    'confidence_level': row['confs'],
                    'mode_of_inheritance': row['mois'],
                    'mode_of_pathogenicity': row['mops'],
                    'penetrance': row['pens'],
                    'name': row['names'],
                    'chrom': row['chroms'],
                    'start_37': row['starts_37'],
                    'end_37': row['ends_37'],
                    'start_38': row['starts_38'],
                    'end_38': row['ends_38'],
                    'type': row['types'],
                    'variant_type': row['var_types'],
                    'required_overlap': row['overlaps'],
                    'haploinsufficiency': row['haplos'],
                    'triplosensitivity': row['triplos'],
                    'justification': row['reasons']}

                info_dict['regions'].append(region_dict)

        return info_dict
