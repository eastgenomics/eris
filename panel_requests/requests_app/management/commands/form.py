#!usr/bin/env python

"""
Given a clinical indication (CI) code, acquires the panel currently
associated with that CI in the database and uses it to populate an Excel
request form.

Generic and example usages can be found in the README.

python manage.py form --help
python manage.py form 20221201 user R149.1
"""
# TODO: dealing with empty Panel or Gene records have yet been looked at

import pandas as pd
import collections

import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment

from django.core.management.base import BaseCommand

from panel_requests.requests_app.models import (
    Panel,
    Gene,
    Confidence,
    Penetrance,
    ModeOfInheritance,
    ModeOfPathogenicity,
    PanelGene,
    Transcript,
    Haploinsufficiency,
    Triplosensitivity,
    RequiredOverlap,
    VariantType,
    Region,
    ClinicalIndication,
    ClinicalIndicationPanel,
)


class Command(BaseCommand):
    help = (
        "Given a clinical indication (CI) code, acquires the panel "
        "currently associated with that CI in the database and uses it "
        "to populate an Excel request form."
    )

    def _get_panel_records(self, ci_code: str) -> list:
        """Retrieve Panel records from the database where the panel is
        associated with the specified CI code, and where that panel is
        currently in use for the specified clinical indication.

        Should return two records, one for each genome build.

        args:
            ci_code [str]: supplied at command line

        returns:
            panel_records [list of dicts]: each dict is one panel record
        """

        associated_panel_ids = ClinicalIndicationPanel.objects.filter(
            clinical_indication_id__code=ci_code, current=True
        ).values_list("panel_id", flat=True)

        panel_records: list[Panel] = []

        for panel_id in associated_panel_ids:
            panel_records.append(Panel.objects.get(id=panel_id))

        return panel_records

    def retrieve_panel_entities(self, panel_records: list):
        """For each retrieved panel record, retrieve all associated
        gene and region records and combine in a dict. Return a list of
        these dicts.

        args:
            ref_genome [str]: supplied at command line
            panel_records [list of dicts]: list of Panel records

        returns:
            panel_dicts [list of dicts]: formatted data for each panel
        """

        panel_data = []

        for panel in panel_records:
            panel_dict = collections.defaultdict(dict)
            panel_ref_genome = []

            if panel.grch37:
                panel_ref_genome.append("GRCh37")

            if panel.grch38:
                panel_ref_genome.append("GRCh38")

            panel_dict[panel.id] = {
                "name": panel.panel_name,
                "db_id": panel.id,
                "source": panel.panel_source,
                "ext_id": panel.external_id,
                "ext_version": panel.panel_version,
                "ref_genome": ", ".join(panel_ref_genome),
                "genes": self._get_gene_records(panel.id),
                "regions": self._get_region_records(panel.id),
            }

            panel_data.append(panel_dict)

        return panel_data

    def _get_gene_records(self, id):
        """Retrieve all gene records associated with a panel record.

        args:
            panel_dict [dict]: information about a single panel record

        returns:
            gene_data [list of dicts]: each element is info on one gene
        """

        gene_data = []

        panel_genes = PanelGene.objects.filter(panel_id=id).values()

        if not panel_genes:
            return []

        for record in panel_genes:
            gene_id = record["gene_id"]
            confidence = Confidence.objects.get(
                id=record["confidence_id"]
            ).confidence_level
            penetrance = Penetrance.objects.get(
                id=record["penetrance_id"]
            ).penetrance
            moi = ModeOfInheritance.objects.get(
                id=record["moi_id"]
            ).mode_of_inheritance
            mop = ModeOfPathogenicity.objects.get(
                id=record["mop_id"]
            ).mode_of_pathogenicity
            gene_instance = Gene.objects.get(id=gene_id)
            justification = record["justification"]

            transcript = Transcript.objects.filter(
                gene_id=gene_id
            ).values_list("transcript", flat=True)

            gene_data.append(
                {
                    "transcript": list(transcript) if transcript else None,
                    "hgnc": gene_instance.hgnc_id,
                    "gene_symbol": gene_instance.gene_symbol,
                    "conf": confidence,
                    "moi": moi,
                    "mop": mop,
                    "pen": penetrance,
                    "panel_gene_justification": justification,
                }
            )

        return gene_data

    def _get_region_records(self, id):
        """Retrieve all region records associated with a panel record.

        args:
            panel_dict [dict]: information about a single panel record

        returns:
            region_data [list of dicts]: each element is info on one region
        """

        region_data = []

        panel_regions = Region.objects.filter(panel_id=id).values()

        if not panel_regions:
            return []

        for record in panel_regions:
            # metadata
            confidence = Confidence.objects.get(
                id=record["confidence_id"]
            ).confidence_level
            penetrance = Penetrance.objects.get(
                id=record["penetrance_id"]
            ).penetrance
            moi = ModeOfInheritance.objects.get(
                id=record["moi_id"]
            ).mode_of_inheritance
            mop = ModeOfPathogenicity.objects.get(
                id=record["mop_id"]
            ).mode_of_pathogenicity
            variant_type = VariantType.objects.get(
                id=record["vartype_id"]
            ).variant_type
            haplo = Haploinsufficiency.objects.get(
                id=record["haplo_id"]
            ).haploinsufficiency
            triplo = Triplosensitivity.objects.get(
                id=record["triplo_id"]
            ).triplosensitivity
            overlap = RequiredOverlap.objects.get(
                id=record["overlap_id"]
            ).required_overlap

            name = record["name"]
            chromosome = record["chrom"]
            start_37 = record["start_37"]
            end_37 = record["end_37"]
            start_38 = record["start_38"]
            end_38 = record["end_38"]
            region_type = record["type"]

            justification = record["justification"]

            region_dict = {
                "conf": confidence,
                "moi": moi,
                "mop": mop,
                "pen": penetrance,
                "name": name,
                "chrom": chromosome,
                "start_37": start_37,
                "end_37": end_37,
                "start_38": start_38,
                "end_38": end_38,
                "type": region_type,
                "var_type": variant_type,
                "haplo": haplo,
                "triplo": triplo,
                "overlap": overlap,
                "panel_region_justification": justification,
            }

            region_data.append(region_dict)

        return region_data

    def _create_generic_df(self, req_date, requester, ci_code):
        """Create a header dataframe of general request information.

        args:
            req_date [str]: supplied at command line
            requester [str]: supplied at command line
            ci_code [str]: supplied at command line

        returns:
            generic_df [pandas dataframe]
        """

        base_df = pd.DataFrame(
            {
                "fields": [
                    "Request Date",
                    "Requested By",
                    "Clinical Indication",
                    "Form Generated",
                ],
                "data": [
                    req_date,
                    requester,
                    ci_code,
                    str(dt.datetime.today()),
                ],
            }
        )

        generic_df = base_df.set_index("fields")

        return generic_df

    def _create_panel_df(self, panel_data: list) -> pd.DataFrame:
        """Create a dataframe of information about panels currently
        associated with the specified clinical indication.

        args:
            panel_dicts [list of dicts]: each dict is one panel

        returns:
            panel_df [pandas dataframe]
        """

        # dataframe columns are lists of data elements

        names = []
        sources = []
        ids = []
        versions = []

        for panel in panel_data:
            for _, panel_dict in panel.items():
                names.append(panel_dict["name"])
                sources.append(panel_dict["source"])
                ids.append(panel_dict["ext_id"])
                versions.append(panel_dict["ext_version"])

        # create the df

        panel_df = pd.DataFrame(
            {
                "Current Panels": names,
                "Panel Source": sources,
                "External ID": ids,
                "External Version": versions,
            }
        )

        return panel_df

    def _create_gene_df(self, panel_data: list) -> pd.DataFrame:
        """Create a dataframe listing all genes currently covered by
        the specified CI's panels.

        args:
            panel_dicts [list of dicts]: each dict is one panel

        returns:
            gene_df [pandas dataframe]
        """

        # initialise df columns

        gene_symbols = []
        hgncs = []
        panel_ids = []
        panel_gene_justification = []
        transcripts = []

        confs = []
        pens = []
        mops = []
        mois = []

        # populate columns

        for panel in panel_data:
            for _, panel_dict in panel.items():
                for gene in panel_dict.get("genes", []):
                    gene_symbols.append(gene["gene_symbol"])
                    hgncs.append(gene["hgnc"])
                    panel_ids.append(
                        panel_dict["db_id"]
                    )  # associated panel id in db
                    panel_gene_justification.append(
                        gene["panel_gene_justification"]
                    )

                    transcripts.append(", ".join(gene.get("transcript", [])))

                    confs.append(gene["conf"])
                    pens.append(gene["pen"])
                    mops.append(gene["mop"])
                    mois.append(gene["moi"])

        gene_df = pd.DataFrame(
            {
                "Gene Symbol": gene_symbols,
                "HGNC ID": hgncs,
                "Panel ID in Database": panel_ids,
                "PanelGene Justification": panel_gene_justification,
                "Transcript": transcripts,
                "Confidence": confs,
                "Penetrance": pens,
                "MOP": mops,
                "MOI": mois,
            }
        )

        print(gene_df.shape)

        return gene_df

    def _create_region_df(self, panel_data):
        """Create a dataframe listing all regions currently covered by
        the specified CI's panels.

        args:
            panel_dicts [list of dicts]: each dict is one panel

        returns:
            region_df [pandas datafra
            build = panel["ref_genome"]["reference_build"]me]
        """

        # initialise df columns

        names = []
        chroms = []
        starts_37 = []
        ends_37 = []
        starts_38 = []
        ends_38 = []
        panel_region_justification = []
        confs = []
        pens = []
        mops = []
        mois = []
        haplos = []
        triplos = []
        types = []
        var_types = []
        overlaps = []

        for panel in panel_data:
            for _, panel_dict in panel.items():
                for region in panel_dict.get("regions", []):
                    names.append(region["name"])
                    chroms.append(region["chrom"])
                    panel_region_justification.append(
                        region["panel_region_justification"]
                    )
                    confs.append(region["conf"])
                    pens.append(region["pen"])
                    mops.append(region["mop"])
                    mois.append(region["moi"])
                    types.append(region["type"])
                    var_types.append(region["var_type"])
                    haplos.append(region["haplo"])
                    triplos.append(region["triplo"])
                    overlaps.append(region["overlap"])

                    starts_37.append(region["start_37"])
                    ends_37.append(region["end_37"])
                    starts_38.append(region["start_38"])
                    ends_38.append(region["end_38"])

        region_df = pd.DataFrame(
            {
                "Region Name": names,
                "Chromosome": chroms,
                "Start (GRCh37)": starts_37,
                "End (GRCh37)": ends_37,
                "Start (GRCh38)": starts_38,
                "End (GRCh38)": ends_38,
                "Justification": panel_region_justification,
                "Confidence": confs,
                "Penetrance": pens,
                "MOP": mops,
                "MOI": mois,
                "Type": types,
                "Variant Type": var_types,
                "Haploinsufficiency": haplos,
                "Triplosensitivity": triplos,
                "Overlap": overlaps,
            }
        )

        return region_df

    def create_blank_dfs(self):
        """For cases where a clinical indication has no associated
        current panels, create blank dfs for genes and regions to be
        added to.

        returns:
            panel_df [pandas dataframe]
            gene_df [pandas dataframe]
            region_df [pandas dataframe]
        """

        panel_df = pd.DataFrame(
            {
                "Current panels": [
                    "None currently associated with this clinical "
                    "indication"
                ],
                "Panel source": [""],
                "External ID": [""],
                "External version": [""],
            }
        )

        gene_df = pd.DataFrame(
            {
                "Gene symbol": ["None currently in panel"],
                "HGNC ID": [""],
                "Gene justification": [""],
                "Transcript": [""],
                "Transcript justification": [""],
                "Confidence": [""],
                "Penetrance": [""],
                "MOP": [""],
                "MOI": [""],
            }
        )

        region_df = pd.DataFrame(
            {
                "Region name": ["None currently in panel"],
                "Chromosome": [""],
                "Start (GRCh37)": [""],
                "End (GRCh37)": [""],
                "Start (GRCh38)": [""],
                "End (GRCh38)": [""],
                "Justification": [""],
                "Confidence": [""],
                "Penetrance": [""],
                "MOP": [""],
                "MOI": [""],
                "Type": [""],
                "Variant type": [""],
                "Haploinsufficiency": [""],
                "Triplosensitivity": [""],
                "Overlap": [""],
            }
        )

        return panel_df, gene_df, region_df

    def write_data(self, fp, generic_df, panel_df, gene_df, region_df):
        """Construct the request form as an excel file.

        args:
            fp [str]: to write request form to
            generic_df [pandas dataframe]: general info about request
            panel_df [pandas dataframe]: list of current panels for this CI
            gene_df [pandas dataframe]: all genes covered by these panels
            region_df [pandas dataframe]: all regions covered by these panels

        returns:
            cell_ranges [list]: strs of cell ranges (e.g. 'A1:A7')
        """

        writer = pd.ExcelWriter(fp, engine="xlsxwriter")

        row = 0
        col = 0

        # holds rows of table headers
        header_rows = {}

        # write generic df in cell A1, increment row by df length

        header_rows["general"] = row + 1

        generic_df.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=row,
            startcol=col,
            header=False,
        )

        for df_row in generic_df.iterrows():
            row += 1

        row += 1  # spacer between dataframes

        # write panel df, get row of header, increment row by df length

        header_rows["panel"] = row + 1

        panel_df.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=row,
            startcol=col,
            index=False,
        )

        for df_row in panel_df.iterrows():
            row += 1

        row += 2

        # write gene df, get row of header, increment row by df length

        header_rows["gene"] = row + 1

        gene_df.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=row,
            startcol=col,
            index=False,
        )

        for df_row in gene_df.iterrows():
            row += 1

        row += 2

        # write region df, get row of header

        header_rows["region"] = row + 1

        region_df.to_excel(
            writer,
            sheet_name="Sheet1",
            startrow=row,
            startcol=col,
            index=False,
        )

        # identify the last populated row

        for df_row in region_df.iterrows():
            row += 1

        header_rows["final"] = row + 7

        writer.close()

        return header_rows

    def format_excel(self, excel_file, header_rows):
        """Visually format an existing excel file. Apply a style to
        dataframe header cells, and set column widths to autofit data.

        args:
            excel_file [str]: path to file
            cell_ranges [list]: strs (e.g. 'A1:A7') of cell ranges for styling
        """

        # load in the excel file

        wb = load_workbook(filename=excel_file)
        ws = wb["Sheet1"]

        final_row = header_rows["final"]

        # define a style for a default font type and size

        normal_font = NamedStyle(name="normal_font")

        normal_font.font = Font(name="Arial", size=10)

        normal_font.alignment = Alignment(horizontal="left", vertical="center")

        wb.add_named_style(normal_font)

        # define a style to highlight column headings

        col_headings = NamedStyle(name="col_headings")

        col_headings.font = Font(name="Arial", size=10, bold=True)

        col_headings.alignment = Alignment(
            horizontal="left", vertical="center"
        )

        col_headings.fill = PatternFill(
            fill_type="solid",
            start_color="00C0C0C0",  # light grey
            end_color="00C0C0C0",
        )

        wb.add_named_style(col_headings)

        # define a style to highlight the end of the form

        highlight = NamedStyle(name="highlight")

        highlight.font = Font(name="Arial", size=10, bold=True)

        highlight.alignment = Alignment(horizontal="left", vertical="center")

        highlight.fill = PatternFill(
            fill_type="solid",
            start_color="00FFCC00",
            end_color="00FFCC00",  # yellow
        )

        wb.add_named_style(highlight)

        # apply the default font to all populated cells

        for col in "ABCDEFGHIJKLMNOPQ":
            for row in range(1, final_row + 1):
                ws[f"{col}{row}"].style = "normal_font"

        # apply the heading style to the index of the generic df

        for row in range(1, 5):
            ws[f"A{row}"].style = "col_headings"

        # apply the heading style to the headers of the other dfs

        for header in "panel", "gene", "region":
            row = header_rows[header]

            if header == "panel":
                cols = "ABCD"

            elif header == "gene":
                cols = "ABCDEFGHI"

            elif header == "region":
                cols = "ABCDEFGHIJKLMNOP"

            for col in cols:
                ws[f"{col}{row}"].style = "col_headings"

        # add form instructions in the final rows

        ws[f"A{final_row - 4}"] = (
            "Each row should contain data for one "
            "gene or region - add or remove rows as required."
        )

        ws[f"A{final_row - 3}"] = "Please leave 1 blank row between tables."

        ws[f"A{final_row - 2}"] = (
            "Please do not change any headings, or "
            "the order of the columns."
        )

        ws[f"A{final_row}"] = "END OF FORM"
        ws[f"B{final_row}"] = "Please do not enter anything below this row."

        for row in range(final_row - 4, final_row + 1):
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].style = "highlight"

        # set all columns to be 5cm wide

        for col in "ABCDEFGHIJKLMNOPQ":
            ws.column_dimensions[col].width = 25.5

        wb.save(filename=excel_file)

    def _validate_args(self, req_date: str, ci_code: str) -> None:
        try:
            dt.datetime.strptime(req_date, "%d%m%Y")
        except ValueError or TypeError:
            raise ValueError("Incorrect date format, should be ddmmyyyy")

        all_cis = set(
            ClinicalIndication.objects.all().values_list("code", flat=True)
        )

        if ci_code not in all_cis:
            raise ValueError("CI code not found in database")

    def add_arguments(self, parser):
        """Define the CI to generate a request form for"""

        parser.add_argument(
            "date",
            help="Date the request was submitted e.g. 15062023 (ddmmyyyy)",
        )

        parser.add_argument(
            "requester",
            help="Person who submitted the request",
        )

        parser.add_argument(
            "ci_code",
            help="The CI code to create request form for (e.g. R149.1)",
        )

    def handle(self, *args, **kwargs):
        """Execute the functions to create a request form"""

        print(kwargs)  # TODO: remove

        # read in arguments (all are required)
        request_date = kwargs.get("date")
        requester = kwargs.get("requester")
        ci_code = kwargs.get("ci_code")

        self._validate_args(request_date, ci_code)

        output_filename = (
            f"request_form_{request_date}_{ci_code}_{requester}.xlsx"
        )

        # construct header df of general info about the request

        generic_df = self._create_generic_df(request_date, requester, ci_code)

        # generate blank dataframes

        panel_df = pd.DataFrame()
        gene_df = pd.DataFrame()
        region_df = pd.DataFrame()

        panels_empty, genes_empty, regions_empty = self.create_blank_dfs()

        # retrieve records of panels currently linked to that CI code

        panel_records = self._get_panel_records(ci_code)

        if panel_records:
            # create panel, gene and region dataframes

            panels_data = self.retrieve_panel_entities(panel_records)

            if not panels_data:
                raise ValueError("No Panel record data found for this CI")

            panel_df = self._create_panel_df(panels_data)
            region_df = self._create_region_df(panels_data)
            gene_df = self._create_gene_df(panels_data)

        # construct the request form and print the filename
        if len(panel_df) == 0:
            panel_df = panels_empty

        if len(gene_df) == 0:
            gene_df = genes_empty

        if len(region_df) == 0:
            region_df = regions_empty

        headers = self.write_data(
            output_filename, generic_df, panel_df, gene_df, region_df
        )

        # apply formatting to the created file

        self.format_excel(output_filename, headers)

        print(f"Request form created: {output_filename}")
