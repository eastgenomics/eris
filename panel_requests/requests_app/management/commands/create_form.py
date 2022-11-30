#!usr/bin/env python

"""
Given a clinical indication (CI) code, acquires the panel currently
associated with that CI in the database and uses it to populate an Excel
request form.

Generic and example usages can be found in the README.
"""


import pandas as pd
import xlsxwriter

from . import functions_hgnc

from copy import copy
from datetime import datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment

from django.core.management.base import BaseCommand

from panel_requests.requests_app.models import (
    ReferenceGenome,
    Panel,
    ClinicalIndication,
    ClinicalIndicationPanel,
    Hgnc,
    Gene,
    Confidence,
    Penetrance,
    ModeOfInheritance,
    ModeOfPathogenicity,
    PanelGene,
    Transcript,
    PanelGeneTranscript,
    Haploinsufficiency,
    Triplosensitivity,
    RequiredOverlap,
    VariantType,
    Region,
    PanelRegion)


class Command(BaseCommand):
    help = "Given a clinical indication (CI) code, acquires the panel " \
        "currently associated with that CI in the database and uses it " \
        "to populate an Excel request form."

    def add_arguments(self, parser):
        """ Define the CI to generate a request form for """

        parser.add_argument(
            "req_date",
            nargs=1,
            help="Date the request was submitted",)

        parser.add_argument(
            "requester",
            nargs=1,
            help="Person who submitted the request",)

        parser.add_argument(
            "ci_code",
            nargs=1,
            help="The CI code to create request form for (e.g. R149.1)",)

        parser.add_argument(
            "hgnc_dump",
            nargs=1,
            help="Name of HGNC dump text file")

    def get_panel_records(self, ci_code):
        """ Retrieve Panel records from the database where the panel is
        associated with the specified CI code, and where that panel is
        currently in use for the specified clinical indication.

        Should return two records, one for each genome build.

        args:
            ci_code [str]: supplied at command line

        returns:
            panel_records [list of dicts]: each dict is one panel record
        """

        panel_records = Panel.objects.filter(
            clinicalindicationpanel__clinical_indication_id__code=ci_code,
            clinicalindicationpanel__current=True).values()

        return panel_records

    def retrieve_panel_entities(self, panel_records):
        """ For each retrieved panel record, retrieve all associated
        gene and region records and combine in a dict. Return a list of
        these dicts.

        args:
            ref_genome [str]: supplied at command line
            panel_records [list of dicts]: list of Panel records

        returns:
            panel_dicts [list of dicts]: formatted data for each panel
        """

        panel_dicts = []

        for panel_dict in panel_records:

            panel_genome = ReferenceGenome.objects.filter(
                panel=panel_dict['id']).values()[0]

            panel_genes = self.get_gene_records(panel_dict)
            panel_regions = self.get_region_records(panel_dict)

            panel_data = {
                'name': panel_dict['panel_name'],
                'int_id': panel_dict['id'],
                'source': panel_dict['panel_source'],
                'ext_id': panel_dict['external_id'],
                'ext_version': panel_dict['panel_version'],
                'ref_genome': panel_genome,
                'genes': panel_genes,
                'regions': panel_regions}

            panel_dicts.append(panel_data)

        return panel_dicts

    def get_gene_records(self, panel_dict):
        """ Retrieve all gene records associated with a panel record.

        args:
            panel_dict [dict]: information about a single panel record

        returns:
            gene_data [list of dicts]: each element is info on one gene
        """

        gene_data = []

        panel_genes = PanelGene.objects.filter(
            panel_id=panel_dict['id']).values()

        if len(panel_genes) > 0:

            for gene in panel_genes:

                transcript = Transcript.objects.filter(
                    panelgenetranscript__panel_gene_id=gene['id']
                    ).values()[0]['refseq_id']

                transcript_reason = PanelGeneTranscript.objects.filter(
                    panel_gene_id=gene['id']
                    ).values()[0]['justification']

                hgnc = Hgnc.objects.filter(
                    gene__panelgene__id=gene['id']
                    ).values()[0]['id']

                confidence = Confidence.objects.filter(
                    panelgene__id=gene['id']
                    ).values()[0]['confidence_level']

                moi = ModeOfInheritance.objects.filter(
                    panelgene__id=gene['id']
                    ).values()[0]['mode_of_inheritance']

                mop = ModeOfPathogenicity.objects.filter(
                    panelgene__id=gene['id']
                    ).values()[0]['mode_of_pathogenicity']

                penetrance = Penetrance.objects.filter(
                    panelgene__id=gene['id']
                    ).values()[0]['penetrance']

                gene_dict = {
                    'transcript': transcript,
                    'hgnc': hgnc,
                    'conf': confidence,
                    'moi': moi,
                    'mop': mop,
                    'pen': penetrance,
                    'gene_reason': gene['justification'],
                    'trans_reason': transcript_reason}

                gene_data.append(gene_dict)

        return gene_data

    def get_region_records(self, panel_dict):
        """ Retrieve all region records associated with a panel record.

        args:
            panel_dict [dict]: information about a single panel record

        returns:
            region_data [list of dicts]: each element is info on one region
        """

        region_data = []

        panel_regions = PanelRegion.objects.filter(
            panel_id=panel_dict['id']).values()

        if len(panel_regions) > 0:

            for region in panel_regions:

                confidence = Confidence.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['confidence_level']

                moi = ModeOfInheritance.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['mode_of_inheritance']

                mop = ModeOfPathogenicity.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['mode_of_pathogenicity']

                penetrance = Penetrance.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['penetrance']

                name = Region.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['name']

                chromosome = Region.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['chrom']

                start = Region.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['start']

                end = Region.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['end']

                region_type = Region.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['type']

                variant_type = VariantType.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['variant_type']

                haplo = Haploinsufficiency.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['haploinsufficiency']

                triplo = Triplosensitivity.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['triplosensitivity']

                overlap = RequiredOverlap.objects.filter(
                    panelregion__id=region['id']
                    ).values()[0]['required_overlap']

                region_dict = {
                    'conf': confidence,
                    'moi': moi,
                    'mop': mop,
                    'pen': penetrance,
                    'name': name,
                    'chrom': chromosome,
                    'start': start,
                    'end': end,
                    'type': region_type,
                    'var_type': variant_type,
                    'haplo': haplo,
                    'triplo': triplo,
                    'overlap': overlap,
                    'reason': region['justification']}

                region_data.append(region_dict)

        return region_data

    def create_generic_df(self, req_date, requester, ci_code):
        """ Create a header dataframe of general request information.

        args:
            req_date [str]: supplied at command line
            requester [str]: supplied at command line
            ci_code [str]: supplied at command line

        returns:
            generic_df [pandas dataframe]
        """

        base_df = pd.DataFrame({
            'fields': [
                'Request date',
                'Requested by',
                'Clinical indication',
                'Form generated'],
            'data': [
                req_date,
                requester,
                ci_code,
                str(dt.today())]})

        generic_df = base_df.set_index('fields')

        return generic_df

    def create_panel_df(self, panel_dicts):
        """ Create a dataframe of information about panels currently
        associated with the specified clinical indication.

        args:
            panel_dicts [list of dicts]: each dict is one panel

        returns:
            panel_df [pandas dataframe]
        """

        # dataframe columns are lists of data elements

        names = []
        sources = []
        ids = []
        versions = []

        for panel in panel_dicts:
            if panel['ext_id'] not in ids:

                names.append(panel['name'])
                sources.append(panel['source'])
                ids.append(panel['ext_id'])
                versions.append(panel['ext_version'])

        # create the df

        panel_df = pd.DataFrame({
            'Current panels': names,
            'Panel source': sources,
            'External ID': ids,
            'External version': versions})

        return panel_df

    def create_gene_df(self, hgnc_df, panel_dicts):
        """ Create a dataframe listing all genes currently covered by
        the specified CI's panels.

        args:
            panel_dicts [list of dicts]: each dict is one panel

        returns:
            gene_df [pandas dataframe]
        """

        # initialise df columns

        gene_symbols = []
        hgncs = []
        gene_reasons = []
        transcripts = []
        trans_reasons = []
        confs = []
        pens = []
        mops = []
        mois = []

        # populate columns

        for panel in panel_dicts:
            for gene in panel['genes']:
                if gene['hgnc'] not in hgncs:

                    hgncs.append(gene['hgnc'])
                    gene_reasons.append(gene['gene_reason'])
                    transcripts.append(gene['transcript'])
                    trans_reasons.append(gene['trans_reason'])
                    confs.append(gene['conf'])
                    pens.append(gene['pen'])
                    mops.append(gene['mop'])
                    mois.append(gene['moi'])

        # get current gene symbols for HGNC ids

        for hgnc_id in hgncs:

            symbol = functions_hgnc.get_symbol_from_hgnc(hgnc_df, hgnc_id)

            if symbol:

                gene_symbols.append(symbol)

            else:

                gene_symbols.append('')  # else list is the wrong length
                print(f'Problem with HGNC number: {hgnc_id}')

        # create df

        gene_df = pd.DataFrame({
            'Gene symbol': gene_symbols,
            'HGNC ID': hgncs,
            'Gene justification': gene_reasons,
            'Transcript': transcripts,
            'Transcript justification': trans_reasons,
            'Confidence': confs,
            'Penetrance': pens,
            'MOP': mops,
            'MOI': mois})

        return gene_df

    def create_region_df(self, panel_dicts):
        """ Create a dataframe listing all regions currently covered by
        the specified CI's panels.

        args:
            panel_dicts [list of dicts]: each dict is one panel

        returns:
            region_df [pandas dataframe]
        """

        # initialise df columns

        names = []
        chroms = []
        starts_37 = []
        ends_37 = []
        starts_38 = []
        ends_38 = []
        reasons = []
        confs = []
        pens = []
        mops = []
        mois = []
        haplos = []
        triplos = []
        types = []
        var_types = []
        overlaps = []

        for panel in panel_dicts:
            build = panel['ref_genome']['reference_build']

            for region in panel['regions']:

                # if region not in column yet, add with whichever build coords
                if region['name'] not in names:

                    names.append(region['name'])
                    chroms.append(region['chrom'])
                    reasons.append(region['reason'])
                    confs.append(region['conf'])
                    pens.append(region['pen'])
                    mops.append(region['mop'])
                    mois.append(region['moi'])
                    types.append(region['type'])
                    var_types.append(region['var_type'])
                    haplos.append(region['haplo'])
                    triplos.append(region['triplo'])
                    overlaps.append(region['overlap'])

                    if build == 'GRCh37':
                        starts_37.append(region['start'])
                        ends_37.append(region['end'])
                        starts_38.append('placeholder')
                        ends_38.append('placeholder')

                    elif build == 'GRCh38':
                        starts_38.append(region['start'])
                        ends_38.append(region['end'])
                        starts_37.append('placeholder')
                        ends_37.append('placeholder')

                # there should be another panel with the other build's coords
                elif region['name'] in names:

                    index = names.index(region['name'])

                    if build == 'GRCh37':
                        starts_37[index] = region['start']
                        ends_37[index] = region['end']

                    elif build == 'GRCh38':
                        starts_38[index] = region['start']
                        ends_38[index] = region['end']

        # create df

        region_df = pd.DataFrame({
            'Region name': names,
            'Chromosome': chroms,
            'Start (GRCh37)': starts_37,
            'End (GRCh37)': ends_37,
            'Start (GRCh38)': starts_38,
            'End (GRCh38)': ends_38,
            'Justification': reasons,
            'Confidence': confs,
            'Penetrance': pens,
            'MOP': mops,
            'MOI': mois,
            'Type': types,
            'Variant type': var_types,
            'Haploinsufficiency': haplos,
            'Triplosensitivity': triplos,
            'Overlap': overlaps})

        return region_df

    def create_blank_dfs(self):
        """ For cases where a clinical indication has no associated
        current panels, create blank dfs for genes and regions to be
        added to.

        returns:
            panel_df [pandas dataframe]
            gene_df [pandas dataframe]
            region_df [pandas dataframe]
        """

        panel_df = pd.DataFrame({
           'Current panels': ['None currently associated with this clinical '
                'indication'],
            'Panel source': [''],
            'External ID': [''],
            'External version': ['']})

        gene_df = pd.DataFrame({
            'Gene symbol': ['None currently in panel'],
            'HGNC ID': [''],
            'Gene justification': [''],
            'Transcript': [''],
            'Transcript justification': [''],
            'Confidence': [''],
            'Penetrance': [''],
            'MOP': [''],
            'MOI': ['']})

        region_df = pd.DataFrame({
            'Region name': ['None currently in panel'],
            'Chromosome': [''],
            'Start (GRCh37)': [''],
            'End (GRCh37)': [''],
            'Start (GRCh38)': [''],
            'End (GRCh38)': [''],
            'Justification': [''],
            'Confidence': [''],
            'Penetrance': [''],
            'MOP': [''],
            'MOI': [''],
            'Type': [''],
            'Variant type': [''],
            'Haploinsufficiency': [''],
            'Triplosensitivity': [''],
            'Overlap': ['']})

        return panel_df, gene_df, region_df

    def write_data(self, fp, generic_df, panel_df, gene_df, region_df):
        """ Construct the request form as an excel file.

        args:
            fp [str]: to write request form to
            generic_df [pandas dataframe]: general info about request
            panel_df [pandas dataframe]: list of current panels for this CI
            gene_df [pandas dataframe]: all genes covered by these panels
            region_df [pandas dataframe]: all regions covered by these panels

        returns:
            cell_ranges [list]: strs of cell ranges (e.g. 'A1:A7')
        """

        writer = pd.ExcelWriter(fp, engine='xlsxwriter')

        row = 0
        col = 0

        # holds rows of table headers
        header_rows = {}

        # write generic df in cell A1, increment row by df length

        header_rows['general'] = row + 1

        generic_df.to_excel(
            writer,
            sheet_name='Sheet1',
            startrow=row,
            startcol=col,
            header=False)

        for df_row in generic_df.iterrows():
            row += 1

        row += 1  # spacer between dataframes

        # write panel df, get row of header, increment row by df length

        header_rows['panel'] = row + 1

        panel_df.to_excel(
            writer,
            sheet_name='Sheet1',
            startrow=row,
            startcol=col,
            index=False)

        for df_row in panel_df.iterrows():
            row += 1

        row += 2

        # write gene df, get row of header, increment row by df length

        header_rows['gene'] = row + 1

        gene_df.to_excel(
            writer,
            sheet_name='Sheet1',
            startrow=row,
            startcol=col,
            index=False)

        for df_row in gene_df.iterrows():
            row += 1

        row += 2

        # write region df, get row of header

        header_rows['region'] = row + 1

        region_df.to_excel(
            writer,
            sheet_name='Sheet1',
            startrow=row,
            startcol=col,
            index=False)

        # identify the last populated row

        for df_row in region_df.iterrows():
            row += 1

        header_rows['final'] = row + 7

        writer.save()

        return header_rows

    def format_excel(self, excel_file, header_rows):
        """ Visually format an existing excel file. Apply a style to
        dataframe header cells, and set column widths to autofit data.

        args:
            excel_file [str]: path to file
            cell_ranges [list]: strs (e.g. 'A1:A7') of cell ranges for styling
        """

        # load in the excel file

        wb = load_workbook(filename=excel_file)
        ws = wb['Sheet1']

        final_row = header_rows['final']

        # define a style for a default font type and size

        normal_font = NamedStyle(name="normal_font")

        normal_font.font = Font(name='Arial', size=10)

        normal_font.alignment = Alignment(
            horizontal='left', vertical='center')

        wb.add_named_style(normal_font)

        # define a style to highlight column headings

        col_headings = NamedStyle(name="col_headings")

        col_headings.font = Font(name='Arial', size=10, bold=True)

        col_headings.alignment = Alignment(
            horizontal='left', vertical='center')

        col_headings.fill = PatternFill(
            fill_type='solid',
            start_color='00C0C0C0',  # light grey
            end_color='00C0C0C0')

        wb.add_named_style(col_headings)

        # define a style to highlight the end of the form

        highlight = NamedStyle(name="highlight")

        highlight.font = Font(name='Arial', size=10, bold=True)

        highlight.alignment = Alignment(
            horizontal='left', vertical='center')

        highlight.fill = PatternFill(
            fill_type='solid',
            start_color='00FFCC00',  # yellow
            end_color='00FFCC00')

        wb.add_named_style(highlight)

        # apply the default font to all populated cells

        for col in 'ABCDEFGHIJKLMNOPQ':
            for row in range(1, final_row + 1):

                ws[f'{col}{row}'].style = 'normal_font'

        # apply the heading style to the index of the generic df

        for row in range(1, 5):

            ws[f'A{row}'].style = 'col_headings'

        # apply the heading style to the headers of the other dfs

        for header in 'panel', 'gene', 'region':

            row = header_rows[header]

            if header == 'panel':
                cols = 'ABCD'

            elif header == 'gene':
                cols = 'ABCDEFGHI'

            elif header == 'region':
                cols = 'ABCDEFGHIJKLMNOP'

            for col in cols:
                ws[f'{col}{row}'].style = 'col_headings'

        # add form instructions in the final rows

        ws[f'A{final_row - 4}'] = 'Each row should contain data for one ' \
            'gene or region - add or remove rows as required.'

        ws[f'A{final_row - 3}'] = 'Please leave 1 blank row between tables.'

        ws[f'A{final_row - 2}'] = 'Please do not change any headings, or ' \
            'the order of the columns.'

        ws[f'A{final_row}'] = 'END OF FORM'
        ws[f'B{final_row}'] = 'Please do not enter anything below this row.'

        for row in range(final_row - 4, final_row + 1):
            for col in ['A', 'B', 'C']:

                ws[f'{col}{row}'].style = 'highlight'

        # set all columns to be 5cm wide

        for col in 'ABCDEFGHIJKLMNOPQ':

            ws.column_dimensions[col].width = 25.5

        wb.save(filename=excel_file)

    def handle(self, *args, **kwargs):
        """ Execute the functions to create a request form """

        # read in arguments (all are required)

        if kwargs['req_date'] and \
            kwargs['requester'] and \
            kwargs['ci_code'] and \
            kwargs['hgnc_dump']:

            req_date = kwargs['req_date'][0]
            requester = kwargs['requester'][0]
            ci_code = kwargs['ci_code'][0]
            hgnc_dump = kwargs['hgnc_dump'][0]

            file = f'request_form_{req_date}_{ci_code}_{requester}.xlsx'

            # construct header df of general info about the request

            generic_df = self.create_generic_df(
                req_date,
                requester,
                ci_code)

            # generate blank dataframes

            panel_df = pd.DataFrame()
            gene_df = pd.DataFrame()
            region_df = pd.DataFrame()

            panels_empty, genes_empty, regions_empty = self.create_blank_dfs()

            # retrieve records of panels currently linked to that CI code

            panel_records = self.get_panel_records(ci_code)

            if panel_records:

                # create panel, gene and region dataframes

                panel_dicts = self.retrieve_panel_entities(panel_records)

                hgnc_df = functions_hgnc.import_hgnc_dump(hgnc_dump)
                hgnc_df = functions_hgnc.rename_columns(hgnc_df)

                panel_df = self.create_panel_df(panel_dicts)
                region_df = self.create_region_df(panel_dicts)
                gene_df = self.create_gene_df(hgnc_df, panel_dicts)

            # construct the request form and print the filename

            if len(panel_df) == 0:
                panel_df = panels_empty

            if len(gene_df) == 0:
                gene_df = genes_empty

            if len(region_df) == 0:
                region_df = regions_empty

            headers = self.write_data(
                file,
                generic_df,
                panel_df,
                gene_df,
                region_df)

            # apply formatting to the created file

            self.format_excel(file, headers)

            print(f'Request form created: {file}')

        else:
            print("The following arguments are required, and must be "
                "specified in this order: request_date, requester, ci_code, "
                "hgnc_dump file.\nValid values for ref_genome are GRCh37 "
                "and GRCh38.")
