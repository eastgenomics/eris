#!usr/bin/env python

import os
import datetime as dt
import pandas as pd
from openpyxl import load_workbook


class FormParser:
    def __init__(self, filepath: str):
        self.filepath = self.validate_filepath(filepath)

    def validate_filepath(self, filepath: str) -> bool:
        """Check that the filepath is valid.

        args:
            filepath [str]: path to request form
        """
        if os.path.exists(filepath):
            return True

        if not filepath.endswith(".xlsx"):
            raise ValueError(f"Filepath {filepath} should end with .xlsx")

        raise ValueError(f"Filepath {filepath} does not exist.")

    def get_form_data(self, fp: str):
        """Pull in data from the request form file and parse into 4
        pandas DFs - general request info, panels info, genes, and
        regions.

        args:
            fp [str]: path to request form

        returns:
            general_info [dict]
            panel_df [pandas df]
            gene_df [pandas df]
            region_df [pandas df]
        """

        wb = load_workbook(filename=fp)
        ws = wb.active

        # identify row boundaries of each data table
        # this assumes there is -exactly- one blank row between each table

        row_count = 1
        p_start, g_start, r_start = None, None, None

        for row in ws.iter_rows(max_col=1, values_only=True):
            if row[0] == "Current Panels":
                p_start = row_count + 1

            elif row[0] == "Gene Symbol":
                g_start = row_count + 1
                p_end = row_count - 2

            elif row[0] == "Region Name":
                r_start = row_count + 1
                g_end = row_count - 2

            elif row[0] == "END OF FORM":
                r_end = row_count - 6
                break

            row_count += 1

        if not p_start or not g_start or not r_start:
            raise ValueError("Issue with form parsing. Check form formatting.")

        # create dict of general request info

        general_info = {
            "req_date": ws["B1"].value,
            "requester": ws["B2"].value,
            "ci": ws["B3"].value,
            "form_created_date": ws["B4"].value,
        }

        # create df from panels info

        panel_df = pd.DataFrame(
            {
                "names": [cell[0].value for cell in ws[f"A{p_start}":f"A{p_end}"]],
                "sources": [cell[0].value for cell in ws[f"B{p_start}":f"B{p_end}"]],
                "ext_ids": [cell[0].value for cell in ws[f"C{p_start}":f"C{p_end}"]],
                "ext_versions": [
                    cell[0].value for cell in ws[f"D{p_start}":f"D{p_end}"]
                ],
                "panel_id_in_database": [
                    cell[0].value for cell in ws[f"E{p_start}":f"E{p_end}"]
                ],
            }
        )

        # create df from genes info

        gene_df = pd.DataFrame(
            {
                "gene_symbols": [
                    cell[0].value for cell in ws[f"A{g_start}":f"A{g_end}"]
                ],
                "hgncs": [cell[0].value for cell in ws[f"B{g_start}":f"B{g_end}"]],
                "reasons": [cell[0].value for cell in ws[f"C{g_start}":f"C{g_end}"]],
                "transcripts": [
                    cell[0].value for cell in ws[f"D{g_start}":f"D{g_end}"]
                ],
                "confs": [cell[0].value for cell in ws[f"E{g_start}":f"F{g_end}"]],
                "pens": [cell[0].value for cell in ws[f"F{g_start}":f"G{g_end}"]],
                "mops": [cell[0].value for cell in ws[f"G{g_start}":f"H{g_end}"]],
                "mois": [cell[0].value for cell in ws[f"H{g_start}":f"I{g_end}"]],
            }
        )

        # create df from regions info

        region_df = pd.DataFrame(
            {
                "names": [cell[0].value for cell in ws[f"A{r_start}":f"A{r_end}"]],
                "chroms": [cell[0].value for cell in ws[f"B{r_start}":f"B{r_end}"]],
                "starts_37": [cell[0].value for cell in ws[f"C{r_start}":f"C{r_end}"]],
                "ends_37": [cell[0].value for cell in ws[f"D{r_start}":f"D{r_end}"]],
                "starts_38": [cell[0].value for cell in ws[f"E{r_start}":f"E{r_end}"]],
                "ends_38": [cell[0].value for cell in ws[f"F{r_start}":f"F{r_end}"]],
                "reasons": [cell[0].value for cell in ws[f"G{r_start}":f"G{r_end}"]],
                "confs": [cell[0].value for cell in ws[f"H{r_start}":f"H{r_end}"]],
                "pens": [cell[0].value for cell in ws[f"I{r_start}":f"I{r_end}"]],
                "mops": [cell[0].value for cell in ws[f"J{r_start}":f"J{r_end}"]],
                "mois": [cell[0].value for cell in ws[f"K{r_start}":f"K{r_end}"]],
                "types": [cell[0].value for cell in ws[f"L{r_start}":f"L{r_end}"]],
                "var_types": [cell[0].value for cell in ws[f"M{r_start}":f"M{r_end}"]],
                "haplos": [cell[0].value for cell in ws[f"N{r_start}":f"N{r_end}"]],
                "triplos": [cell[0].value for cell in ws[f"O{r_start}":f"O{r_end}"]],
                "overlaps": [cell[0].value for cell in ws[f"P{r_start}":f"P{r_end}"]],
            }
        )

        return general_info, panel_df, gene_df, region_df

    def _validate_form_data(self, ci: str, req_date: str) -> bool:
        if not ci:
            raise ValueError("CI not found in request form.")

        try:
            dt.datetime.strptime(req_date, "%d%m%Y")
        except ValueError:
            raise ValueError("Request date not in correct format.")

        return True

    def setup_output_dict(self, info: dict, panels: pd.DataFrame) -> dict:
        """Initialise a dict to hold relevant panel information.

        args:
            ci [str]: clinical indication to link the panel to
            req_date [str]: date of new panel request
            panel_df [pandas df]: PA panels new panel is based on

        returns:
            info_dict [dict]: initial dict of core panel info
        """
        ci = info.get("ci")
        request_date = info.get("req_date")
        requester = info.get("requester")

        self._validate_form_data(ci, request_date)

        info_dict = {
            "ci": ci,
            # required for panel record creation
            "panel_name": panels["names"][0],
            "panel_source": f"ExcelForm_{request_date}_{requester}",
            "panel_version": panels["ext_versions"][0],
            "external_id": panels["ext_ids"][0],
            "panel_id_in_database": panels["panel_id_in_database"][0],
            # required for panel record creation
            "genes": [],
            "regions": [],
        }

        return info_dict

    def parse_genes(self, info_dict: dict, gene_df: pd.DataFrame) -> dict:
        """Iterate over every gene in the panel and retrieve the data
        needed to populate panel_gene and associated models. Only use
        genes with 'confidence_level' == '3'; i.e. 'green' genes.

        args:
            info_dict [dict]: holds data needed to populate db models
            gene_df [pandas df]: genes to include in new panel

        returns:
            info_dict [dict]: updated with gene data
        """

        for _, row in gene_df.iterrows():
            if row["hgncs"]:
                gene_dict = {
                    "hgnc_id": row["hgncs"],
                    "gene_symbol": row["gene_symbols"],
                    "gene_justification": row["reasons"],
                    "confidence_level": row["confs"],
                    "mode_of_inheritance": row["mois"],
                    "mode_of_pathogenicity": row["mops"],
                    "penetrance": row["pens"],
                    "transcript": row["transcripts"],
                }

                info_dict["genes"].append(gene_dict)

        return info_dict

    def parse_regions(self, info_dict: dict, region_df: pd.DataFrame) -> dict:
        """Iterate over every region in the panel and retrieve the data
        needed to populate panel_region and associated models. Only use
        regions with 'confidence_level' == '3'; i.e. 'green' regions.

        args:
            info_dict [dict]: holds data needed to populate db models
            region_df [pandas df]: regions to include in new panel

        returns:
            info_dict [dict]: updated with region data
        """

        for _, row in region_df.iterrows():
            if row["chroms"]:
                region_dict = {
                    "confidence_level": row["confs"],
                    "mode_of_inheritance": row["mois"],
                    "mode_of_pathogenicity": row["mops"],
                    "penetrance": row["pens"],
                    "name": row["names"],
                    "chrom": row["chroms"],
                    "start_37": row["starts_37"],
                    "end_37": row["ends_37"],
                    "start_38": row["starts_38"],
                    "end_38": row["ends_38"],
                    "type": row["types"],
                    "variant_type": row["var_types"],
                    "required_overlap": row["overlaps"],
                    "haploinsufficiency": row["haplos"],
                    "triplosensitivity": row["triplos"],
                    "justification": row["reasons"],
                }

                info_dict["regions"].append(region_dict)

        return info_dict
